import json
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import ttk as tkttk
from tkinter import messagebox
from crud import get_db_connection
import datetime
import pandas as pd
import os
import subprocess
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Side, Border

from ass import *
from GS import get_preferred_class


import AE

class Ranking:
    def __init__(self, root):
        self.root = root
        self.root.title("Student Ranking")
        root.attributes("-topmost", False)

        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_width = 400
        window_height = 250
        window_x = (screen_width - window_width) // 2
        window_y = (screen_height - window_height) // 2
        self.root.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")
        self.root.resizable(False, False)

        frame_padding = (20, 20)
        label_font = ("Helvetica", 12)
        entry_padding_x = 10
        entry_padding_y = 5
        button_padding_x = 10
        button_padding_y = 5

        self.frame = ttk.Frame(self.root, padding=frame_padding)
        self.frame.pack(fill=ttk.BOTH, expand=ttk.TRUE)

        ttk.Label(self.frame, text="Enter Exam Title:", font=label_font).grid(row=0, column=0, padx=entry_padding_x, pady=entry_padding_y, sticky=ttk.W)
        self.exam_title_var = ttk.StringVar()
        self.exam_title_entry = ttk.Entry(self.frame, textvariable=self.exam_title_var, width=50)
        self.exam_title_entry.grid(row=0, column=1, padx=entry_padding_x, pady=entry_padding_y, sticky=ttk.W)

        ttk.Label(self.frame, text="Select Year:", font=label_font).grid(row=1, column=0, padx=entry_padding_x, pady=entry_padding_y, sticky=ttk.W)
        self.year_var = ttk.StringVar()
        self.year_select = ttk.Combobox(self.frame, textvariable=self.year_var, values=self.generate_years())
        self.year_select.grid(row=1, column=1, padx=entry_padding_x, pady=entry_padding_y, sticky=ttk.W)
        self.year_select.bind("<<ComboboxSelected>>")

        ttk.Label(self.frame, text="Select Class:", font=label_font).grid(row=2, column=0, padx=entry_padding_x, pady=entry_padding_y, sticky=ttk.W)
        self.class_var = ttk.StringVar()
        self.class_select = ttk.Combobox(self.frame, textvariable=self.class_var)
        self.class_select.grid(row=2, column=1, padx=entry_padding_x, pady=entry_padding_y, sticky=ttk.W)
        self.class_select.bind("<<ComboboxSelected>>")
        self.populate_class_names()

        ttk.Label(self.frame, text="Select Semester:", font=label_font).grid(row=3, column=0, padx=entry_padding_x, pady=entry_padding_y, sticky=ttk.W)
        self.semester_var = ttk.StringVar()
        self.semester_select = ttk.Combobox(self.frame, textvariable=self.semester_var, values=["1", "2", "3"])
        self.semester_select.grid(row=3, column=1, padx=entry_padding_x, pady=entry_padding_y, sticky=ttk.W)
        self.semester_select.bind("<<ComboboxSelected>>")

        button_frame = ttk.Frame(self.frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=button_padding_y, sticky=ttk.EW)

        self.download_button = ttk.Button(button_frame, text="Download", command=self.download_ranking, bootstyle="success")
        self.download_button.pack(side=ttk.LEFT, padx=button_padding_x)

        self.reset_button = ttk.Button(button_frame, text="Reset", command=self.reset_fields, bootstyle="danger")
        self.reset_button.pack(side=ttk.LEFT, padx=button_padding_x)

        self.exam_title_var.set(self.getTitle())
        
        
        
                # Set default values based on preferred year and semester
        preferred_values = AE.get_preferred_year_semester()
        if "year" in preferred_values:
            self.year_var.set(preferred_values["year"])
        else:
            self.year_var.set(AE.generate_years()[0])  # Set to the most recent year if no preference

        if "semester" in preferred_values:
            self.semester_var.set(preferred_values["semester"])
        else:
            self.semester_var.set("1")  # Set to "1" if no preference

    def generate_years(self):
        current_year = datetime.datetime.now().year
        start_year = 1991
        end_year = current_year + 1
        return [f"{year}/{year + 1}" for year in range(end_year, start_year - 1, -1)]

    def populate_class_names(self):
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, class_name FROM class_name")
        self.class_names = {row["class_name"]: row["id"] for row in cursor.fetchall()}
        conn.close()
        
        class_names = list(self.class_names.keys())

        # Get the preferred class filter
        preferred_class = get_preferred_class()

        if preferred_class:
            # Filter class names that match the preferred class pattern
            filtered_class_names = [
                class_name for class_name in class_names 
                if preferred_class.lower() in class_name.lower()
            ]
            self.class_select["values"] = filtered_class_names
        else:
            self.class_select["values"] = class_names



    def download_ranking(self):
        exam_title = self.exam_title_var.get()
        year = self.year_var.get()
        class_name = self.class_var.get()
        semester = self.semester_var.get()

        if not all([exam_title, year, class_name, semester]):
            messagebox.showwarning("Warning", "Please select all options before downloading.")
            return

        self.setTitle(exam_title)

        filters = []
        if year:
            filters.append(f"Year: {year}")
        if class_name:
            filters.append(f"Class: {class_name}")
        if semester:
            filters.append(f"Semester: {semester}")

        filter_message = ", ".join(filters)

        class_id = self.class_names.get(class_name) if class_name else None
        self.getRanking(exam_title, class_id, year, semester)

    def reset_fields(self):
        self.exam_title_var.set("")
        self.year_var.set("")
        self.class_var.set("")
        self.semester_var.set("")

    def setTitle(self, title):
        HOME_DIR = os.path.expanduser('~')
        APP_DIR = os.path.join(HOME_DIR, 'SHSStudentReportSystem')
        title_file = os.path.join(APP_DIR, 'title.json')

        with open(title_file, 'w') as file:
            json.dump({"title": title}, file)

    def getTitle(self):
        HOME_DIR = os.path.expanduser('~')
        APP_DIR = os.path.join(HOME_DIR, 'SHSStudentReportSystem')
        title_file = os.path.join(APP_DIR, 'title.json')

        if os.path.exists(title_file):
            with open(title_file, 'r') as file:
                data = json.load(file)
                return data.get("title", "")
        return ""

    def get_school_details(self):
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM school_details LIMIT 1")
        school_details = cursor.fetchone()
        conn.close()
        return school_details
    
    def getProgrammeName(self, class_id):
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # First, get the programme_id from the assessment table using the class_id
        query1 = '''
            SELECT programme_id
            FROM assessment
            WHERE class_id = ?
            LIMIT 1
        '''
        cursor.execute(query1, (class_id,))
        result1 = cursor.fetchone()
        
        if result1:
            programme_id = result1['programme_id']
            
            # Now, get the programme_name from the programme table using the programme_id
            query2 = '''
                SELECT programme_name
                FROM programme
                WHERE id = ?
            '''
            cursor.execute(query2, (programme_id,))
            result2 = cursor.fetchone()
            
            conn.close()
            return result2['programme_name'] if result2 else "Unknown Programme"
        
        conn.close()
        return "Unknown Programme"
    
    
    def getRanking(self, exam_title, class_id=None, year=None, semester_id=None):
        # Set directory paths
        HOME_DIR = os.path.expanduser('~')
        APP_DIR = os.path.join(HOME_DIR, 'SHSStudentReportSystem')
        DATABASE_FILENAME = 'shs.db'
        DATABASE_FILE = os.path.join(APP_DIR, DATABASE_FILENAME)
        REPORT_DIR = os.path.join(HOME_DIR, 'Documents', 'RANKING')

        # Ensure the report directory exists
        os.makedirs(REPORT_DIR, exist_ok=True)

        conn = get_db_connection()
        cursor = conn.cursor()

        query = '''
            SELECT ca.student_id, st.name as student_name, sub.subject_name, sub.short_name, ca.year, ca.class_score, ca.exam_score, ca.total_score, ca.grade, ca.number_equivalence, ca.remarks, ca.isCore, ca.teacher_initial_letters, ca.rank,
            RANK() OVER (PARTITION BY ca.student_id ORDER BY ca.total_score DESC) AS rank 
            FROM computed_assessment ca
            JOIN student st ON ca.student_id = st.student_id
            JOIN subject sub ON ca.subject_id = sub.id
            WHERE 1=1
        '''
        params = []

        if class_id:
            query += ' AND ca.class_id = ?'
            params.append(class_id)
        if year:
            query += ' AND ca.year = ?'
            params.append(year)
        if semester_id:
            query += ' AND ca.semester_id = ?'
            params.append(semester_id)

        query += ' ORDER BY ca.student_id, ca.total_score DESC'

        cursor.execute(query, tuple(params))
        rankings = cursor.fetchall()
        print(rankings)
        conn.close()

        # Function to convert position to ordinal
        def ordinal(n):
            if 10 <= n % 100 <= 20:
                suffix = 'th'
            else:
                suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
            return str(n) + suffix

        # Process and arrange rankings similar to the provided image
        student_data = {}
        core_subjects = set()
        elective_subjects = set()
        for row in rankings:
            student_id = row['student_id']
            subject_short_name = row['short_name']
            
            if student_id not in student_data:
                student_data[student_id] = {
                    'student_name': row['student_name'],
                    'subjects': {},
                    'total': 0,
                    'average': 0,
                    'aggregate': get_student_aggregate(row['student_id'], class_id, semester_id, year),
                    'position': 0
                }
            student_data[student_id]['subjects'][subject_short_name] = row['total_score']
            student_data[student_id]['total'] += row['total_score']
            if row['isCore'] == 1:
                core_subjects.add(subject_short_name)
            else:
                elective_subjects.add(subject_short_name)

        for student in student_data.values():
            student['average'] = round(student['total'] / len(student['subjects']), 2)

        sorted_students = sorted(student_data.values(), key=lambda x: x['total'], reverse=True)

        for i, student in enumerate(sorted_students, 1):
            student['position'] = ordinal(i)

        # Arrange core subjects starting with Math, English, Science
        core_subject_order = ["MATH", "ENG", "SCI"]
        ordered_core_subjects = sorted(core_subjects, key=lambda sub: next((i for i, prefix in enumerate(core_subject_order) if prefix in sub), len(core_subject_order)))
        
        subjects = ordered_core_subjects + sorted(core_subjects - set(ordered_core_subjects)) + sorted(elective_subjects)

        # Prepare the data for export
        total_columns = len(subjects) + 6  # Number of subjects plus 6 additional columns
        export_data = []
        for student in sorted_students:
            student_row = [student['student_name']] + [student['subjects'].get(subject, 0) for subject in subjects] + [student['total'], student['average'], student['aggregate'], student['position']]
            export_data.append(student_row)

        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Rankings"

        # Define bold border
        bold_border = Border(left=Side(style='thick'),
                            right=Side(style='thick'),
                            top=Side(style='thick'),
                            bottom=Side(style='thick'))

        # Add school details and header
        school_details = self.get_school_details()
        programme_name = self.getProgrammeName(class_id)
        
        school_name = school_details['school_name'] if school_details else "Unknown School"
        description = f"{school_name} - {exam_title}\n"
        if year:
            description += f" Year: {year}\n"
        if class_id:
            class_name = [name for name, id_ in self.class_names.items() if id_ == class_id][0]
            description += f" Class : {class_name+" "}\n"
        if semester_id:
            description += f"Semester : {semester_id}"

        ws.merge_cells('A1:' + get_column_letter(total_columns) + '1')
        ws.merge_cells('A2:' + get_column_letter(total_columns) + '2')
        ws.merge_cells('A3:' + get_column_letter(total_columns) + '3')

        ws['A1'] = school_name
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = exam_title+": "+programme_name
        ws['A2'].font = Font(size=12, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        classYearSemster=class_name+"  Year: "+year+"  Semester:"+semester_id
        ws['A3'] = classYearSemster
        ws['A3'].font = Font(size=12, bold=True)
        ws['A3'].alignment = Alignment(horizontal='center')

        # Add the header
        header = ["NO.", "NAME OF STUDENT"] + subjects + ["TOTAL", "AVERAGE", "AGGREGATE", "POSITION"]
        ws.append(header)

        # Add the data
        for i, student_row in enumerate(export_data, 1):
            row = [i] + student_row
            ws.append(row)

        # Style the header
        for cell in ws[4]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)
            cell.border = bold_border

        ws.row_dimensions[4].height = 70

        # Style the data cells and center text starting from column 2
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(header)):
            for cell in row:
                cell.border = bold_border
                if cell.column > 2:  # Center text starting from column 2
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        ws.column_dimensions['A'].width = 5  # Adjust width of column A
        ws.column_dimensions['I'].width = 15
        # Concatenate the file name
        file_name = classYearSemster.replace(" ", "_").replace("/", "_").replace(":", "_") + ".xlsx"
        # Save the file
        file_path = os.path.join(REPORT_DIR, file_name)
        wb.save(file_path)
        wb.close()
        print(f"Data exported to {file_path}")

        # Auto open the folder
        if os.name == 'nt':  # For Windows
            os.startfile(REPORT_DIR)
        elif os.name == 'posix':  # For macOS and Linux
            subprocess.call(['open', REPORT_DIR]) if sys.platform == 'darwin' else subprocess.call(['xdg-open', REPORT_DIR])


 
            

if __name__ == "__main__":
    root = ttk.Window(themename="darkly")
    app = Ranking(root)
    root.mainloop()
